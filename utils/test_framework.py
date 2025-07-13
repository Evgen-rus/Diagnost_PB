"""
Модуль для тестирования нейроконсультанта.

Предоставляет инструменты для проведения тестирования бота
по схеме "вопрос-ответ" с автоматическим сбором метрик и
оценкой результатов.
"""
import time
import asyncio
import os
import json
from typing import List, Dict, Any, Optional
from datetime import datetime

from utils.token_counter import token_counter
from utils.logger import logger, get_user_logger
from utils.database import log_test_result, export_test_results_to_csv
from prompts import LEARNING_ASSISTANT_PROMPT

def detect_file_format(file_path: str) -> str:
    """
    Определяет формат файла по расширению.
    
    Args:
        file_path: Путь к файлу
    
    Returns:
        Формат файла: 'xlsx' или 'csv'
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xlsx':
        return 'xlsx'
    return 'csv'  # По умолчанию CSV

def save_to_excel(file_path: str, data: List[List[Any]], headers: List[str]) -> None:
    """
    Сохраняет данные в Excel-файл с форматированием.
    
    Args:
        file_path: Путь к файлу Excel
        data: Данные для записи (список строк)
        headers: Заголовки столбцов
    """
    try:
        import xlsxwriter
        
        # Создаем директорию для файла, если она не существует
        os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else '.', exist_ok=True)
        
        # Создаем новую книгу Excel и лист
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet('Результаты тестирования')
        
        # Создаем форматы для ячеек
        header_format = workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#D9EAD3',
            'border': 1
        })
        
        cell_format = workbook.add_format({
            'align': 'left',
            'valign': 'vcenter',
            'text_wrap': True,
            'border': 1
        })
        
        number_format = workbook.add_format({
            'align': 'right',
            'valign': 'vcenter',
            'border': 1
        })
        
        date_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'num_format': 'yyyy-mm-dd'
        })
        
        time_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'num_format': 'hh:mm:ss'
        })
        
        # Записываем заголовки
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Записываем данные
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, cell_value in enumerate(row_data):
                # Применяем форматирование в зависимости от типа данных и столбца
                if col_idx == 0:  # Столбец "Дата"
                    worksheet.write(row_idx, col_idx, cell_value, date_format)
                elif col_idx == 1:  # Столбец "Время"
                    worksheet.write(row_idx, col_idx, cell_value, time_format)
                elif isinstance(cell_value, (int, float)):
                    worksheet.write(row_idx, col_idx, cell_value, number_format)
                else:
                    worksheet.write(row_idx, col_idx, cell_value, cell_format)
        
        # Устанавливаем ширину столбцов с учетом новых столбцов
        column_widths = {
            0: 12,  # Дата
            1: 10,  # Время
            2: 30,  # Вопрос
            3: 60,  # Ответ
            4: 10,  # Оценка
            5: 10,  # Токены
            6: 30,  # Комментарий
            7: 12,  # Время генерации
            8: 12,  # Цена
            9: 50,  # Чанк 1
            10: 50, # Чанк 2
            11: 50  # Чанк 3
        }
        
        for col, width in column_widths.items():
            if col < len(headers):
                worksheet.set_column(col, col, width)
        
        # Включаем фильтры
        worksheet.autofilter(0, 0, len(data), len(headers) - 1)
        
        # Закрываем книгу
        workbook.close()
        
        logger.info(f"Данные успешно сохранены в Excel-файл: {file_path}")
        
    except ImportError:
        logger.error("Не удалось импортировать модуль xlsxwriter. Установите его с помощью pip install xlsxwriter")
        raise
    except Exception as e:
        logger.error(f"Ошибка при сохранении в Excel: {str(e)}")
        raise

class TestTable:
    """
    Класс для проведения тестирования нейроконсультанта.
    
    Позволяет проводить тесты с заданными вопросами, измерять
    метрики (токены, время, стоимость) и собирать оценки от тестировщиков.
    """
    
    def __init__(self, output_csv: Optional[str] = None):
        """
        Инициализирует таблицу тестирования.
        
        Args:
            output_csv: Путь к файлу для сохранения результатов (CSV или XLSX)
                       Если не указан, результаты сохраняются только в БД.
        """
        self.output_file = output_csv
        self.logger = get_user_logger(user_id=0, operation="test_framework")
        self.logger.info("Инициализация системы тестирования нейроконсультанта")
        
        # Если указан выходной файл, создаем его с заголовками
        if self.output_file:
            # Определяем формат по расширению
            self.file_format = detect_file_format(self.output_file)
            
            # Заголовки для таблицы с добавлением даты и времени
            self.headers = [
                "Дата", "Время", "Вопрос", "Ответ", "Оценка", "Токены", "Комментарий", 
                "Время генерации", "Цена $", "Чанк 1", "Чанк 2", "Чанк 3"
            ]
            
            # Проверяем, существует ли директория для файла
            os.makedirs(os.path.dirname(self.output_file) if os.path.dirname(self.output_file) else '.', exist_ok=True)
            
            # Создаем файл в зависимости от формата
            if self.file_format == 'xlsx':
                if not os.path.isfile(self.output_file):
                    # Создаем пустой Excel файл с заголовками
                    save_to_excel(self.output_file, [], self.headers)
                    self.logger.info(f"Создан Excel-файл для результатов: {self.output_file}")
            else:
                # CSV формат (как было раньше)
                import csv
                
                # Проверяем, существует ли файл
                file_exists = os.path.isfile(self.output_file)
                
                # Создаем файл и записываем заголовки, если он не существует
                with open(self.output_file, mode="a", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    if not file_exists:
                        writer.writerow(self.headers)
                        self.logger.info(f"Создан CSV-файл для результатов: {self.output_file}")
    
    async def run_test(self, question: str, bot_instance, tester_id: int = 0):
        """
        Проводит одиночный тест с указанным вопросом.
        
        Args:
            question: Вопрос для теста
            bot_instance: Экземпляр бота (не используется напрямую)
            tester_id: ID тестировщика (по умолчанию 0)
            
        Returns:
            dict: Результаты теста или None в случае ошибки
        """
        try:
            self.logger.info(f"Начало теста. Вопрос: {question}")
            
            # Замеряем время начала запроса
            start_time = time.time()
            
            # Получаем ответ от модели
            try:
                system_content = LEARNING_ASSISTANT_PROMPT
                
                # Формируем сообщения для запроса
                messages = [{"role": "user", "content": question}]
                
                # Получаем ответ с информацией о чанках, используя новую функцию
                from bot import get_gpt_response_with_chunks
                response, chunks = await get_gpt_response_with_chunks(
                    messages=messages, 
                    system_content=system_content,
                    user_id=tester_id
                )
                
                # Замеряем время окончания запроса
                end_time = time.time()
                generation_time = end_time - start_time
                
                # Получаем данные о токенах и стоимости
                tokens = token_counter.total_tokens
                cost = token_counter.total_cost
                
                # Выводим информацию о результате теста
                print("\n" + "="*70)
                print(f"Вопрос: {question}")
                print("-"*70)
                print(f"Ответ: {response}")
                print("-"*70)
                print(f"Метрики:")
                print(f"  Токены: {tokens}")
                print(f"  Стоимость: ${cost:.6f}")
                print(f"  Время генерации: {generation_time:.2f} сек.")
                if chunks:
                    chunk_ids = [chunk.get('id', 'unknown') if isinstance(chunk, dict) else str(chunk) for chunk in chunks]
                    print(f"  Использованные чанки: {', '.join(chunk_ids)}")
                else:
                    print("  Чанки не использовались")
                print("="*70 + "\n")
                
                # Запрашиваем оценку от тестировщика
                while True:
                    try:
                        score_input = input("Оцените ответ от -2 до +2 (-2, -1, 0, +1, +2): ")
                        score = int(score_input)
                        if -2 <= score <= 2:
                            break
                        print("Оценка должна быть целым числом от -2 до +2!")
                    except ValueError:
                        print("Пожалуйста, введите целое число!")
                
                # Запрашиваем комментарий
                comment = input("Введите комментарий к оценке: ")
                
                # Получаем текущую дату и время для записи в таблицу
                current_datetime = datetime.now()
                current_date = current_datetime.strftime("%Y-%m-%d")
                current_time = current_datetime.strftime("%H:%M:%S")
                
                # Сохраняем результат в БД
                log_test_result(
                    tester_id=tester_id,
                    question=question,
                    answer=response,
                    score=score,
                    tokens=tokens,
                    comment=comment,
                    generation_time=generation_time,
                    cost=cost,
                    chunks=chunks
                )
                
                # Сохраняем результат в файл, если он указан
                if self.output_file:
                    # Формируем данные для 3 столбцов чанков
                    chunk_columns = ["", "", ""]  # По умолчанию пустые
                    if chunks:
                        for i, chunk in enumerate(chunks[:3]):  # Максимум 3 чанка
                            if isinstance(chunk, dict):
                                chunk_text = f"{chunk.get('id', 'unknown')}\n{chunk.get('similarity', 0.0)}\n{chunk.get('content', '')}"
                                chunk_columns[i] = chunk_text
                    
                    row_data = [
                        current_date,  # Дата
                        current_time,  # Время
                        question, 
                        response, 
                        score, 
                        tokens, 
                        comment, 
                        f"{generation_time:.2f}", 
                        f"{cost:.6f}", 
                        chunk_columns[0],  # Чанк 1
                        chunk_columns[1],  # Чанк 2
                        chunk_columns[2]   # Чанк 3
                    ]
                    
                    if self.file_format == 'xlsx':
                        # Для Excel нужно прочитать существующие данные и добавить новую строку
                        try:
                            import xlsxwriter
                            from openpyxl import load_workbook
                            
                            # Проверяем, существует ли файл
                            if os.path.isfile(self.output_file):
                                # Загружаем данные из существующего файла
                                wb = load_workbook(self.output_file)
                                ws = wb.active
                                
                                # Собираем все существующие данные
                                existing_data = []
                                for row in ws.iter_rows(min_row=2, values_only=True):
                                    existing_data.append(list(row))
                                
                                # Добавляем новые данные
                                existing_data.append(row_data)
                                
                                # Сохраняем в новый Excel-файл с форматированием
                                save_to_excel(self.output_file, existing_data, self.headers)
                            else:
                                # Создаем новый файл с одной строкой данных
                                save_to_excel(self.output_file, [row_data], self.headers)
                                
                        except ImportError:
                            self.logger.error("Для работы с Excel требуется установить библиотеки xlsxwriter и openpyxl")
                            print("Для работы с Excel требуется установить библиотеки: pip install xlsxwriter openpyxl")
                            # Сохраняем в CSV как запасной вариант
                            import csv
                            with open(self.output_file.replace('.xlsx', '.csv'), mode="a", newline="", encoding="utf-8") as f:
                                writer = csv.writer(f)
                                writer.writerow(row_data)
                                self.logger.info(f"Результаты сохранены в CSV (запасной вариант): {self.output_file.replace('.xlsx', '.csv')}")
                    else:
                        # CSV формат (как было раньше)
                        import csv
                        with open(self.output_file, mode="a", newline="", encoding="utf-8") as f:
                            writer = csv.writer(f)
                            writer.writerow(row_data)
                
                self.logger.info(f"Тест завершен. Оценка: {score}")
                return {
                    "date": current_date,
                    "time": current_time,
                    "question": question,
                    "answer": response,
                    "score": score,
                    "tokens": tokens,
                    "comment": comment,
                    "generation_time": generation_time,
                    "cost": cost,
                    "chunks": chunks
                }
                
            except Exception as e:
                self.logger.error(f"Ошибка при проведении теста: {str(e)}")
                print(f"Произошла ошибка: {str(e)}")
                return None
                
        except Exception as e:
            self.logger.error(f"Общая ошибка при проведении теста: {str(e)}")
            print(f"Произошла ошибка: {str(e)}")
            return None
    
    async def run_tests_from_file(self, questions_file: str, bot_instance, tester_id: int = 0):
        """
        Проводит серию тестов с вопросами из файла.
        
        Args:
            questions_file: Путь к файлу с вопросами (по одному на строку)
            bot_instance: Экземпляр бота для проведения тестов (не используется напрямую)
            tester_id: ID тестировщика (по умолчанию 0)
        """
        try:
            # Проверяем существование файла
            if not os.path.isfile(questions_file):
                self.logger.error(f"Файл с вопросами не найден: {questions_file}")
                print(f"Ошибка: файл {questions_file} не найден!")
                return
            
            # Загружаем вопросы из файла
            with open(questions_file, "r", encoding="utf-8") as f:
                questions = [line.strip() for line in f if line.strip()]
            
            if not questions:
                self.logger.warning(f"Файл с вопросами пуст: {questions_file}")
                print("Файл с вопросами пуст!")
                return
            
            self.logger.info(f"Загружено {len(questions)} вопросов из файла {questions_file}")
            print(f"Будет выполнено {len(questions)} тестов.\n")
            
            # Проводим тесты последовательно
            results = []
            for i, question in enumerate(questions, 1):
                print(f"\nТест {i}/{len(questions)}")
                result = await self.run_test(question, bot_instance, tester_id)
                if result:
                    results.append(result)
            
            # Выводим статистику
            if results:
                avg_score = sum(r["score"] for r in results) / len(results)
                avg_tokens = sum(r["tokens"] for r in results) / len(results)
                avg_time = sum(r["generation_time"] for r in results) / len(results)
                total_cost = sum(r["cost"] for r in results)
                
                print("\n" + "="*50)
                print(f"ИТОГИ ТЕСТИРОВАНИЯ ({len(results)} тестов):")
                print(f"Средняя оценка: {avg_score:.2f}")
                print(f"Среднее количество токенов: {avg_tokens:.1f}")
                print(f"Среднее время генерации: {avg_time:.2f} сек.")
                print(f"Общая стоимость: ${total_cost:.6f}")
                print("="*50)
                
                self.logger.info(f"Завершено {len(results)} тестов. Средняя оценка: {avg_score:.2f}")
            
            # Подтверждаем сохранение результатов
            if self.output_file:
                print(f"\nРезультаты сохранены в файл: {self.output_file}")
                self.logger.info(f"Результаты тестирования сохранены в {self.output_file}")
            
            return results
                
        except Exception as e:
            self.logger.error(f"Ошибка при проведении тестов из файла: {str(e)}")
            print(f"Произошла ошибка: {str(e)}")
            return []

    @staticmethod
    def export_results(output_file: str = "test_results.csv"):
        """
        Экспортирует результаты тестов из базы данных в файл.
        
        Args:
            output_file: Путь к файлу для экспорта (по умолчанию "test_results.csv")
        """
        try:
            from utils.database import get_test_results
            
            # Получаем результаты из базы данных
            results = get_test_results()
            
            if not results:
                print("Нет результатов тестирования для экспорта.")
                return
            
            # Определяем формат файла
            file_format = detect_file_format(output_file)
            
            if file_format == 'xlsx':
                # Экспорт в Excel
                try:
                    # Подготавливаем данные для Excel
                    headers = [
                        "ID", "Дата", "Время", "Тестировщик ID", "Вопрос", "Ответ", 
                        "Оценка", "Токены", "Комментарий", "Время генерации (сек)", 
                        "Стоимость ($)", "Чанк 1", "Чанк 2", "Чанк 3"
                    ]
                    
                    data = []
                    for result in results:
                        row = [
                            result[0],  # id
                            result[1],  # date
                            result[2],  # time
                            result[3],  # tester_id
                            result[4],  # question
                            result[5],  # answer
                            result[6],  # score
                            result[7],  # tokens
                            result[8],  # comment
                            result[9],  # generation_time
                            result[10], # cost
                            result[11], # chunk 1
                            result[12], # chunk 2
                            result[13]  # chunk 3
                        ]
                        data.append(row)
                    
                    save_to_excel(output_file, data, headers)
                    print(f"Результаты экспортированы в Excel файл: {output_file}")
                    
                except ImportError:
                    print("Для экспорта в Excel требуется установить библиотеку xlsxwriter")
                    # Экспортируем в CSV как запасной вариант
                    csv_file = output_file.replace('.xlsx', '.csv')
                    export_test_results_to_csv(csv_file)
                    print(f"Результаты экспортированы в CSV файл (запасной вариант): {csv_file}")
            else:
                # Экспорт в CSV
                export_test_results_to_csv(output_file)
                print(f"Результаты экспортированы в CSV файл: {output_file}")
                
        except Exception as e:
            print(f"Ошибка при экспорте результатов: {str(e)}")

    @staticmethod
    def export_to_csv(output_file: str = "test_results.csv"):
        """
        Экспортирует результаты тестов из базы данных в CSV файл.
        
        Args:
            output_file: Путь к CSV файлу для экспорта (по умолчанию "test_results.csv")
        """
        TestTable.export_results(output_file) 